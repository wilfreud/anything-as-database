from openpyxl import load_workbook, Workbook
from os import path
from columnar import columnar

# Tables definitions
TABLES = ["USERS", "CONTACTS"]

# Table column definitions
USER_COLUMNS = ["USERNAME", "PASSWORD", "CREATE_AT"]
CONTACTS_COLUMNS = ["FIRSTNAME", "MIDDLENAME", "LASTNAME", "PHONE", "EMAIL", "COMMENT", "CREATED_AT"]
COLUMNS = {} # Associate each table to its columns
COLUMNS[TABLES[0]] = USER_COLUMNS
COLUMNS[TABLES[1]] = CONTACTS_COLUMNS

class ContactManager:
    """
    Contacts Manager Class

    Allows management of contacts for a defined user
    """

    def __init__(self, filepath=None):
        # Set filepath
        self.is_new = filepath is None or filepath == "" or not filepath.endswith("xlsx")
        self.filepath = "db.xlsx" if self.is_new else filepath
        self.is_new = not path.isfile(self.filepath)
        self.wb = load_workbook(self.filepath) if not self.is_new else Workbook()
        self.users = None
        self.contacts = None

        # initialize tables
        self.init_tables()

    def init_tables(self):
        # Remove default sheet
        if self.is_new:
            del self.wb[self.wb.active.title]

        for table in TABLES:
            attr_name = table.lower()
            # Create tables if not existing
            if table not in self.wb.sheetnames:
                print("Creating column: ", table)
                cols = COLUMNS[table]
                setattr(self, attr_name, self.wb.create_sheet(table))
                
                # Set columns
                cols_number = len(cols)
                table_sheet = getattr(self, attr_name).iter_cols(min_row=1, max_col=cols_number, max_row=1)

                for idx, col in enumerate(table_sheet):
                    for cell in col:
                        cell.value = cols[idx]
            else:
                setattr(self, attr_name, self.wb[table])
        
        # Save (always)
        self.save()


    
    def get_tables(self):
        data = []
        for name in self.wb.sheetnames:
            data.append([name])
        
        table = columnar(data, [" TABLES "])
        print(table)

    def get_columns(self, tablename):
        if tablename not in self.wb.sheetnames:
            raise IndexError("Table not found")

        table = getattr(self, tablename.lower())
        if table is None:
            raise KeyError(f"\"{tablename}\" table not found!")

        col = table.iter_cols(min_row=1, max_col=len(COLUMNS[tablename]), max_row=1)
        data = []
        for cell_t in col:
            for cell in cell_t:
                print(cell.value)
            # data.append([cell.value])
        
        # print(columnar(data, [tablename]))
        
    
    def save(self):
        self.wb.save(self.filepath)


manager = ContactManager()
# manager.get_tables();
manager.get_columns(TABLES[0])
